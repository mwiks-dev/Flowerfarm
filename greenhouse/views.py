from django.shortcuts import render
from django.shortcuts import redirect, render, get_object_or_404
from django.core.paginator import Paginator
from django.db.models import Q, Sum
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.cache import never_cache
from django.contrib.auth.views import LoginView
from django.urls import reverse_lazy
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic.edit import UpdateView, CreateView
from django.views.generic import View
from django.contrib.auth.decorators import user_passes_test
from django.contrib.auth import logout
from django.contrib.auth import get_user_model
from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str  # Use force_str here
from django.shortcuts import render, redirect
from django.contrib.auth import login
from django.contrib.auth.forms import SetPasswordForm
from django.contrib.auth.views import LogoutView
from django.db.models.functions import TruncDate
from django import template
from django.utils.dateparse import parse_date


import calendar
from datetime import datetime
import qrcode
from openpyxl import Workbook
from io import BytesIO


from.models import Production, User, RejectedData
from .forms import ProductionForm, RejectedDataForm

# Create your views here.  

#login view
class CustomLoginView(LoginView):
    template_name = 'registration/login.html' 

User = get_user_model()

def custom_activation_view(request, uidb64, token):
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)

        if default_token_generator.check_token(user, token):
            # Create a SetPasswordForm for the user
            if request.method == 'POST':
                form = SetPasswordForm(user, request.POST)
                if form.is_valid():
                    form.save()
                    # Log the user in after setting the password
                    login(request, user)
                    return redirect('custom_activation_success')
            else:
                form = SetPasswordForm(user)

            return render(request, 'custom_activation.html', {'form': form})

    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        pass

    # Activation failed; you can customize this error page
    return render(request, 'activation_error.html')


class UserDetailUpdateView(LoginRequiredMixin, UpdateView):
    model = User  
    template_name = 'update_profile.html'  # Create this template
    fields = ['prof_photo','full_name', 'email','phone_number'] 
    success_url = '/greenhouse/profile/'

    def get_object(self, queryset=None):
        # Retrieve the user object based on the provided primary key (pk)
        return get_object_or_404(self.model, pk=self.kwargs['pk'])

    def form_valid(self, form):
        # Save the updated user details
        user = form.save()
        # You can perform additional actions if needed
        return super().form_valid(form)
    

#index/calendar view function
# @login_required(login_url='/login/')
# def calendar_view(request):
    
#     today = datetime.date.today()
#     year = today.year
#     month = today.month
#     cal = calendar.monthcalendar(year, month)
    
    
#     return render(request, 'index.html', context)
#choice page
@login_required(login_url='/login/')
def choice_page(request):
    current_user = request.user
    username = current_user.full_name

    context = {
        'username':username,
    }

    return render(request, 'choice.html', context)

#user details page
@login_required(login_url='/login/')
def user_details(request):
    current_user = request.user
    profile = User.objects.filter(id=current_user.id).first()

    return render(request,"profile.html",{'profile':profile})

#production form
class ProductionCreateView(CreateView):
    model = Production
    form_class = ProductionForm
    template_name = 'production.html'
    success_url = reverse_lazy('report_choice_page')

    def form_valid(self, form):
        # Associate the user with the profile
        form.instance.user = self.request.user 
        # print(form.cleaned_data['varieties'])
        # self.object = form.save(commit = False)        
        # form.save()
        selected_date_str = self.request.GET.get('selected_date', None)

        if selected_date_str:
            selected_date = datetime.strptime(selected_date_str, "%Y-%m-%d").date()
            form.instance.production_date = selected_date
        return super().form_valid(form)
    
#rejected data form
class RejectedDataCreateView(CreateView):
    model = RejectedData
    form_class = RejectedDataForm
    template_name = 'reject.html'
    success_url = reverse_lazy('report_choice_page')

    def form_valid(self, form):
        form.instance.user = self.request.user

        return super().form_valid(form)

#report choice page
def report_choice_page(request):

    return render(request, 'report_choice.html')
   
#report generation
def production_report(request):
    data = Production.objects.order_by('-production_date')

    # Search functionality
    query = request.GET.get('q')
    if query:
        data = data.filter(Q(varieties__icontains=query) | Q(greenhouse_number__icontains=query))
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if start_date and end_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
        data = data.filter(production_date__range=(start_date, end_date))

    # Aggregations
    total_by_variety = data.values('varieties').annotate(total=Sum('total_number')).order_by('varieties')
    total_by_greenhouse = data.values('greenhouse_number').annotate(total=Sum('total_number')).order_by('greenhouse_number')
    total_by_variety_length = data.values('varieties', 'length').annotate(total=Sum('total_number')).order_by('varieties')
    total_by_greenhouse_length = data.values('greenhouse_number', 'length').annotate(total=Sum('total_number')).order_by('greenhouse_number')

    # Calculating totals
    total_amount_variety = sum(item['total'] for item in total_by_variety)
    total_amount_greenhouse = sum(item['total'] for item in total_by_greenhouse)
    total_amount_variety_length = sum(item['total'] for item in total_by_variety_length)
    total_amount_greenhouse_length = sum(item['total'] for item in total_by_greenhouse_length)

    paginator = Paginator(data, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'page_obj': page_obj,
        'total_amount_variety': total_amount_variety,
        'total_amount_greenhouse': total_amount_greenhouse,
        'total_amount_variety_length': total_amount_variety_length,
        'total_amount_greenhouse_length': total_amount_greenhouse_length,
        'total_by_variety': total_by_variety,
        'total_by_greenhouse': total_by_greenhouse,
        'total_by_variety_length': total_by_variety_length,
        'total_by_greenhouse_length': total_by_greenhouse_length,
        'start_date': start_date,
        'end_date': end_date
    }

    return render(request, 'production_report.html', context)

#rejection report
def rejection_report(request):
    data = RejectedData.objects.order_by('-rejection_date')

    # Search functionality
    query = request.GET.get('q')
    if query:
        data = data.filter(Q(variety__icontains=query) | Q(greenhouse_number__icontains=query))
    paginator = Paginator(data, 10)

    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'rejected_report.html',{'page_obj':page_obj} )

#qr code generation
@method_decorator(csrf_exempt, name='dispatch')
@method_decorator(never_cache, name='dispatch')
class GenerateQRCodeView(View):
    def get(self, request, pk):
        # Retrieve the object from the model based on the primary key (pk)
        prod_data = get_object_or_404(Production, pk=pk)
        
        # Concatenate the fields into a single string
        data_to_encode = f"Date: {prod_data.production_date},  Variety: {prod_data.varieties}, Length: {prod_data.length}, GreenHouse Number: {prod_data.greenhouse_number}, Staff: {prod_data.user}"
        
        # Create a QR code instance
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        
        # Add data to the QR code
        qr.add_data(data_to_encode)
        qr.make(fit=True)
        
        # Create an image from the QR code
        img = qr.make_image(fill_color="black", back_color="white")
        
        # Serve the image as a response
        response = HttpResponse(content_type="image/png")
        img.save(response, "PNG")
        
        return response


def download_excel_report(request):
    # Assuming you're recalculating the data here
    data = Production.objects.all()
    total_by_variety = data.values('varieties').annotate(total=Sum('total_number'))
    total_by_greenhouse = data.values('greenhouse_number').annotate(total=Sum('total_number'))
    total_by_variety_length = data.values('varieties', 'length').annotate(total=Sum('total_number'))
    total_by_greenhouse_length = data.values('greenhouse_number', 'length').annotate(total=Sum('total_number'))

    total_amount_variety = sum(item['total'] for item in total_by_variety)
    total_amount_greenhouse = sum(item['total'] for item in total_by_greenhouse)
    total_amount_variety_length = sum(item['total'] for item in total_by_variety_length)
    total_amount_greenhouse_length = sum(item['total'] for item in total_by_greenhouse_length)

    # Retrieve and parse date filters from the request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if start_date:
        start_date = parse_date(start_date)
    if end_date:
        end_date = parse_date(end_date)

    # Filter data based on dates
    data = Production.objects.all()
    if start_date:
        data = data.filter(production_date__gte=start_date)
    if end_date:
        data = data.filter(production_date__lte=end_date)

    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Production Reports"

    # Add date range information at the top of the sheet
    date_range_info = f"Date Range: {start_date.strftime('%Y-%m-%d') if start_date else 'N/A'} to {end_date.strftime('%Y-%m-%d') if end_date else 'N/A'}"
    ws.append([date_range_info])
    ws.append([])  # Add an empty row for spacing


    # Add headers to the worksheet
    headers = ['Production Date', 'Greenhouse Number', 'Varieties', 'Total Number', 'Length', 'Staff Number']
    ws.append(headers)

    # Add data rows
    for item in data:
        row = [
            item.production_date.strftime('%Y-%m-%d') if item.production_date else '',
            item.greenhouse_number,
            item.varieties,
            item.total_number,
            item.length,
            item.user.staff_number if item.user else ''
        ]
        ws.append(row)

    # Adding data for variety
    ws1 = wb.create_sheet(title='Variety Totals')
    ws1.append(['Variety', 'Total'])
    for item in total_by_variety:
        ws1.append([item['varieties'], item['total']])
    ws1.append(['Sum of Total',total_amount_variety])

    # Adding data for greenhouse in a new sheet
    ws2 = wb.create_sheet(title="Greenhouse Totals")
    ws2.append(['Greenhouse Number', 'Total'])
    for item in total_by_greenhouse:
        ws2.append([item['greenhouse_number'], item['total']])
    ws2.append(['Sum of Total',total_amount_greenhouse])

    # Adding data for variety and length in a new sheet
    ws3 = wb.create_sheet(title="Variety and Length Totals")
    ws3.append(['Variety','Length','Total'])
    for item in total_by_variety_length:
        ws3.append([item['varieties'],item['length'], item['total']])
    ws3.append(['Sum of Total',total_amount_variety_length])


    # Adding data for greenhouse and lengths in a new sheet
    ws4 = wb.create_sheet(title="GreenHouse Number and Length Total")
    ws4.append(['Greenhouse Number','Length','Total'])
    for item in total_by_greenhouse_length:
        ws4.append([item['greenhouse_number'],item['length'],item['total']])
    ws4.append(['Sum of Total',total_amount_greenhouse_length])

    # Save the workbook to a BytesIO object
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="report.xlsx"'
    with BytesIO() as b:
        wb.save(b)
        b.seek(0)
        response.write(b.read())

    return response
class CustomLogoutView(LogoutView):
    next_page = 'login'
# def custom_logout(request):
#     farewell_message = "Goodbye, {}! We hope to see you again soon.".format(request.user.username)
#     logout(request)

#     # Redirect the user to a different page after logout.
#     return render(request, 'registration/logout.html', {'message':farewell_message})
    

