import openpyxl as op
from datetime import datetime
import pandas as pd
from sqlalchemy import create_engine


# file = op.load_workbook('Data.xlsx')

# sheets = file.sheetnames
# # print(sheets)
# sheet = file[sheets[0]]

# # Get all columns headers (assuming they are in the first row)
# columns = [sheet.cell(row=1, column=i).value for i in range(1, sheet.max_column + 1)]

# # Identify the indices of columns that start with 'Date'
# date_column_indices = [i for i, col in enumerate(columns, start=1) if col and col.startswith('Date')]

# # Iterate through each cell in the identified date columns
# for col_idx in date_column_indices:
#     for row in range(2, sheet.max_row + 1):  # Starting from row 2 to skip the header
#         cell = sheet.cell(row=row, column=col_idx)
#         if cell.value is not None:
#             # Check if the cell value is a datetime object
#             if isinstance(cell.value, datetime):
#                 # Format the date as yy/mm/dd and update the cell value
#                 cell.value = cell.value.strftime('%y-%m-%d')
#             else:
#                 # Handle cases where the cell value is not a datetime object
#                 print(f"Non-date value in cell {cell.coordinate}: {cell.value}")


# # Save the modified workbook
# file.save('data1.xlsx')

excel_file = "data1.xlsx"  
data = pd.read_excel(excel_file)

df = pd.DataFrame(data)
# print(df)

# Ignore rows where 'Variety' is 'Total'
df = df[df['Variety'] != 'Total']

# Get all the date and data columns
date_columns = [col for col in df if col.startswith('Date')]
data_columns = [col for col in df if col.startswith('Data')]

# Melt the date columns into a single 'Production Date' column
df_dates = df.melt(id_vars=['Variety', 'GH.', 'Grade'], value_vars=date_columns, var_name='Date_Col', value_name='Production Date')

# Melt the data columns into a single 'Total Number' column
df_data = df.melt(id_vars=['Variety', 'GH.', 'Grade'], value_vars=data_columns, var_name='Data_Col', value_name='Total Number')

# Ensure the order of rows in both DataFrames is the same before concatenating
df_data = df_data.sort_values(by=['Variety', 'GH.', 'Grade', 'Data_Col']).reset_index(drop=True)
df_dates = df_dates.sort_values(by=['Variety', 'GH.', 'Grade', 'Date_Col']).reset_index(drop=True)

# Concatenate the 'Production Date' and 'Total Number' into a single DataFrame
df_final = pd.concat([df_dates['Production Date'], df_data['Total Number']], axis=1)

# Combine 'Variety', 'GH.' and 'Grade' into the final DataFrame
df_final['Variety'] = df_dates['Variety']
df_final['Greenhouse Number'] = df_dates['GH.']
df_final['Length'] = df_dates['Grade']

# Convert the 'Production Date' to the format yyyy-mm-dd
df_final['Production Date'] = pd.to_datetime(df_final['Production Date'], format='%y-%m-%d').dt.strftime('%Y-%m-%d')

# Filter out the rows where 'Production Date' or 'Total Number' is NaN (if required)
df_final = df_final.dropna(subset=['Production Date', 'Total Number'])

# Rename columns to the desired names
df_final.rename(columns={'GH.': 'Greenhouse Number', 'Grade': 'Length'}, inplace=True)

# Reorder the columns
df_final = df_final[['Production Date', 'Greenhouse Number', 'Variety', 'Total Number', 'Length']]

# Convert 'Greenhouse Number' and 'Length' to their proper data types, if needed
df_final['Greenhouse Number'] = df_final['Greenhouse Number'].dropna().astype(int).astype(str)
df_final['Total Number'] = df_final['Total Number'].dropna()
df_final['Length'] = df_final['Length'].dropna().astype(int).astype(str)
df_final['Variety'] = df_final['Variety'].str.title()


# Sort the DataFrame by 'Production Date' and 'Greenhouse Number'
df_final = df_final.sort_values(by=['Production Date', 'Greenhouse Number'])

# Rename columns to the desired names
df_final.rename(columns={
    'Greenhouse Number': 'greenhouse_number', 
    'Variety': 'varieties', 
    'Production Date': 'production_date', 
    'Length': 'length',
    'Total Number': 'total_number'
}, inplace=True)

# The rest of your code remains the same

# df_final.to_csv('new.xlsx')

print(df_final)
print(df_final.info())

def feed_data_to_mysql(df_final, table_name, mysql_url):
    # Create a connection to the MySQL database
    engine = create_engine(mysql_url)
    
    # Feed the data into the MySQL database
    df_final.to_sql(name=table_name, con=engine, if_exists='append', index=False)
    
    print(f"Data has been successfully inserted into the '{table_name}' table.")

# excel_file = pd.read_excel(df_final, index_col=0, engine='openpyxl')

# Define your MySQL connection URL
# Format: mysql+pymysql://<username>:<password>@<host>/<dbname>
mysql_url = 'mysql+pymysql://root:Vodca$4543.@localhost/Flowerfarm'

# Feed the data into the 'mytable' table in the MySQL database
feed_data_to_mysql(df_final, 'greenhouse_production', mysql_url)